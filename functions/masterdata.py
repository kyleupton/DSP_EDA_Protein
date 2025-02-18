import time
import os
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from copy import copy
from IPython.display import clear_output
import itertools as itertools
from collections import Counter as cnt


class MasterData:
    def __init__(self, dataPath):
        # import data from excel workbook
        self.wb = load_workbook(dataPath)
        self.ws = self.wb['Exported dataset']
        self.values = [[y.value for y in x]
                       for x in self.ws[self.ws.calculate_dimension()]]
        self.dropData = False
        self.threshold = False

    def get_data(self, fix_zeros=True, clean_names=True):
        # Convert nested list to a pandas dataFrame and extract
        # expression data with labels
        df = pd.DataFrame(self.values)
        col3 = df.iloc[:, 3].tolist()
        self.targIdx = col3.index('Target name (display name)') + 1
        rowLabels = df.iloc[self.targIdx:, 3]
        # rowLabels = [x.split(' (')[0] for x in rowLabels.values]
        rowLabels = dict(zip([x for x in range(
            self.targIdx, self.targIdx+len(rowLabels))], rowLabels))
        rowLabels
        colLabels = df.iloc[0, 4:]
        colLabels = [x.replace(' | ', '_') for x in colLabels.values]
        colLabels = dict(zip([x for x in range(
            4, 4+len(colLabels))], colLabels))
        colLabels
        self.data = df.iloc[self.targIdx:, 4:].astype(np.float32)
        self.data.rename(index=rowLabels, columns=colLabels, inplace=True)
        if fix_zeros:
            for x in self.data.columns:
                # print(x)
                for y in self.data.index:
                    # print(y)
                    # Subtract 1 from hyb pos and hyb neg values as we
                    # currently don't use RCC files to determine which values
                    # have been changed from 0 to 1 for Neg control.
                    # Changing hyb-pos values may help to ameliorate effects
                    # of changing true 1 values to zero, average decrease in
                    # hyb neg values.
                    if (y == 'HYB-NEG'):
                        self.data.loc[y, x] = self.data.loc[y, x] - 1
                    # elif (y == 'HYB-POS'):
                    #     self.data.loc[y, x] = self.data.loc[y, x] - 1
                    else:
                        if (self.data.loc[y, x] == 1):
                            # print(x,y)
                            self.data.loc[y, x] = 0.0
        self.sampleInfo = pd.DataFrame(df.iloc[0:self.targIdx-1, 4:])
        self.sampleInfo.rename(index=df.iloc[
                               0:self.targIdx-1, 0],
                               columns=colLabels,
                               inplace=True)
        # print('sampleInfo.shape')
        # print(self.sampleInfo.shape)
        # print('data.shape')
        # print(self.data.shape)
        if clean_names:
            removeChars = [' ', '#', '$', '.', ', ', '(', ')', '-', '/', '\\',
                           '__', '__', '__', '__']
            for r in removeChars:
                self.data.columns = [x.replace(
                                     r, '_') for x in self.data.columns]
                self.data.columns = ['X'+x if x[0] in ['0', '1', '2', '3',
                                     '4', '5', '6', '7', '8', '9'] else x
                                     for x in self.data.columns]
                self.sampleInfo.columns = [x.replace(r, '_') for x in
                                           self.sampleInfo.columns]
                self.sampleInfo.columns = ['X'+x if x[0] in ['0', '1', '2',
                                           '3', '4', '5', '6', '7', '8', '9']
                                           else x for x in
                                           self.sampleInfo.columns]
        self.dataOrig = self.data.copy()
        # Log transform data for QC and analysis steps
        self.dataLog1 = np.log2(self.data+1)
        self.probeClass = df.iloc[self.targIdx:, 2]
        self.probeClass.rename(index=rowLabels, inplace=True)
        self.probeClass.rename(index='ProbeClass', inplace=True)
        self.probeClassDict = {
            'Positive': 'A',
            'Negative': 'B',
            'Control': 'C',
            'Endogenous': 'E'
        }
        return self.data.copy(), self.dataLog1.copy(), self.sampleInfo.copy()

    def get_descriptors(self):
        # Extract descriptions for each sample
        nuclei = sampleInfo.loc['AOI nuclei count']
        surfArea = sampleInfo.loc['AOI surface area']
        # print(sampleInfo.shape)

    def add_class_mean(self, df):
        # Add column to data with mean values for each probe (row)
        mean = df.mean(axis=1)
        df = df.assign(mean=mean.values)
        # Add column to data with probe class for each probe
        df = df.assign(probeClass=[self.probeClassDict[v] for v in
                       self.probeClass.values])
        # Extract lists of controls and their values
        self.posCTLs = self.probeClass.index[self.probeClass == 'Positive']\
            .tolist()
        self.negCTLs = self.probeClass.index[self.probeClass == 'Negative']\
            .tolist()
        self.IgCTLs = copy(self.negCTLs)
        try:
            self.IgCTLs.remove('HYB-NEG')
        except ValueError:
            pass
        self.HK = self.probeClass.index[self.probeClass == 'Control']\
            .tolist()
        self.endog = self.probeClass.index[self.probeClass == 'Endogenous']\
            .tolist()
        print('Positive Control count:\t{:d}, {}'.format(len(self.posCTLs),
              self.posCTLs))
        print('Nagative Control count:\t{:d}, {}'.format(len(self.negCTLs),
              self.negCTLs))
        print('Ig Control count:\t{:d}, {}'.format(len(self.IgCTLs),
              self.IgCTLs))
        print('HK Control count:\t{:d}, {}'.format(len(self.HK), self.HK))
        print('Endogenous probe count:\t{:d}, {}'.format(len(self.endog),
              self.endog))
        return df.copy(), self.sampleInfo.copy()

    def drop_AOIs(self, includes, writeOrig=False):
        dropAOIs = [x for x in list(self.data.columns) if (x in includes)]
        # print(dropAOIs)
        if writeOrig:
            self.dataOrig = self.dataOrig.drop(labels=dropAOIs, axis=1,
                                               inplace=True)
        # print(self.dataLog1.shape)
        self.dataLog1.drop(labels=dropAOIs, axis=1, inplace=True)
        # print(self.dataLog1.shape)
        try:
            self.ERCCData.drop(labels=dropAOIs, axis=1, inplace=True)
        except:
            pass
        self.sampleInfo.drop(labels=dropAOIs, axis=1, inplace=True)
        return self.dataLog1.copy(), self.sampleInfo.copy()

    def set_threshold(self, threshold):
        self.threshold = threshold
        # ToDo: Check that all values in master data are also included in
        # threshold dataFrame
        # ToDo: Convert threshold data to 0/1 data if needed

    def drop_probes(self, labels):
        try:
            assert type(labels) == list
        except:
            print('labels need to be a list')
            return False
        if not self.dropData:
            self.dropData = self.dataLog1
        self.dropData.drop(labels=labels)
        return self.dropData.copy()

    def ERCC_norm(self):
        self.ERCCData = self.dataLog1
        try:
            self.ERCCData = self.ERCCData.drop(labels=['mean'], axis=1)
        except:
            pass
        try:
            self.ERCCData = self.ERCCData.drop(labels=['probeClass'], axis=1)
        except:
            pass
        # ERCC normalisation.
        # Divide by individual HYB-POS values, then scale data using the
        # geometric mean of all HYB-POS values. subtract in log space is same
        # as divide in normal space. Add in log space is same as multiply in
        # normal space
        self.ERCCData = self.ERCCData - self.ERCCData.loc['HYB-POS'] +\
            np.mean(self.ERCCData.loc['HYB-POS'])
        # #ToDo: set below threshold values to 0
        # if (self.threshold):
        #     self.ERCCData = self.ERCCData * self.threshold
        # set any negative values following ERCC noirmalisation to 0. These
        # are at or below the limit of detection
        for x in self.ERCCData.columns:
            # print(x)
            for y in self.ERCCData.index:
                if (self.ERCCData.loc[y, x] < 0.0):
                    # print(x,y)
                    self.ERCCData.loc[y, x] = 0.0
# #         zeroFilter = self.ERCCData > 0
#         self.ERCCData = self.ERCCData * (self.ERCCData>0)
        return self.ERCCData.copy()


def check_plate_info(sampleInfo):
    # Check Well, Row, Plate, Col fields in in sampleInfo
    # Number of values should be less than number of samples if info has not
    # been added.
    if (sampleInfo.loc[['Well', 'Row', 'Col', 'Plate']].isnull().any().any()):
        return False
    else:
        return(True)
# ToDo : create function to check if all wells in a plate match by
# surface area


def infer_plate_info(sampleInfo, rootDir, wsFiles):
    # global sampleInfo
    # global configDict
    sampleInfo.loc['AOI surface area'] = \
                   [round(x) for x in sampleInfo.loc['AOI surface area']]
    indexList = list(map(chr, range(ord('A'), ord('H')+1)))
    columnList = [str(n).zfill(2) for n in range(1, 13)]
    validWells = []
    for i in indexList:
        for c in columnList:
            validWells.append(i+c)
    wsList = []
    wellDFs = []  # Set up empty dataframe to be populated with sample names
    for x in wsFiles:
        wsList.append(read_Surf_Areas(os.path.join(rootDir, x.strip()),
                      indexList, columnList))
        wellDFs.append(pd.DataFrame(data='', index=indexList,
                       columns=columnList))
    wsAreaList = []
    allArea = []
    print('wsList')
    print(wsList)
    print(len(wsList))
    for i, ws in enumerate(wsList):
        wsAreaList.append(list(wsList[i].values.flatten()))
        allArea.extend(wsAreaList[i])
    print(allArea)
    allArea = [k for k in allArea if not np.isnan(k)]
    collect = cnt(allArea)
    print(collect)
    # collect = [k for k in collect if not np.isnan(k)]
    print(collect)
    unique = [int(k) for k in collect.keys() if collect[k] == 1]
    nonUnique = [int(k) for k in collect.keys() if collect[k] != 1]
    # unique = [k for k in collect if collect[k] ==1]
    # unique = [k for k in unique if not np.isnan(k)]
    # unique = [int(k) for k in unique]
    # nonUnique = [k for k in collect.keys() if collect[k] !=1]
    # nonUnique = [k for k in unique if not np.isnan(k)]
    # nonUnique = [int(k) for k in unique]
    plates = (wsList[0],)
    SAWellDict = {}
    SAPlateDict = {}
    AOItoWellDict = {}
    AOItoPlateDict = {}
    PlateWellDict = {1: [], 2: []}
    for i, plate in enumerate(wsList):
        for col in plate.columns:
            print(col)
            for row in plate.index:
                print(row)
                # val = int(plate.loc[row, col])
                val = plate.loc[row, col]
                if np.isnan(val):
                    val = 0
                    continue
                val = int(val)
                if ((not val == 0) and (val in unique)):
                    SAWellDict[val] = row + col
                    possibleMatches = sampleInfo.loc[:, sampleInfo.T[
                                      'AOI surface area'] == val].columns
                    if len(possibleMatches == 1):
                        wellDFs[i].loc[row, col] = possibleMatches[0]
                        AOItoWellDict[possibleMatches[0]] = row + col
                    else:
                        print('possibleMatches')
                        print(possibleMatches)
                        print(val)
                        continue
                    if val in wsAreaList[i]:
                        print(val)
                        SAPlateDict[val] = 1
                        PlateWellDict[i+1].append(row+col)
                        AOItoPlateDict[possibleMatches[0]] = i+1
    toLocate = make_locate_list(sampleInfo, AOItoWellDict)
    AOItoPlateDict, AOItoWellDict, PlateWellDict, wellDFs = enter_locations(
        toLocate, validWells, AOItoPlateDict, AOItoWellDict,
        PlateWellDict, wellDFs)
    AOIWell = pd.DataFrame(data=AOItoWellDict.values(),
                           columns=['Well'], index=AOItoWellDict.keys()).T
    AOIRow = pd.DataFrame(data=[x[0] for x in AOItoWellDict.values()],
                          columns=['Row'], index=AOItoWellDict.keys()).T
    AOICol = pd.DataFrame(data=[x[1:] for x in AOItoWellDict.values()],
                          columns=['Col'], index=AOItoWellDict.keys()).T
    AOIPlate = pd.DataFrame(data=AOItoPlateDict.values(),
                            columns=['Plate'], index=AOItoPlateDict.keys()).T
    plateInfo = pd.concat([AOIWell, AOIRow, AOICol, AOIPlate])
    sampleInfo = pd.concat([sampleInfo, plateInfo])
    return(sampleInfo)


def read_Surf_Areas(wsPath, indexList, columnList):
    with open(wsPath, 'r')as f:
        results = []
        record = False
        lines = f.readlines()
        for line in lines:
            line = line.strip()
            if line.startswith('Area '):
                record = True
            elif line.startswith('Totals '):
                record = False
            elif record:
                fields = line.split()
                results.append([int(x) for x in fields[1:]])
    return(pd.DataFrame(results, index=indexList, columns=columnList))


def make_locate_list(sampleInfo, AOItoWellDict):
    toLocate = []
    for smpl in sampleInfo.columns:
        # print(smpl)
        if not (smpl in AOItoWellDict.keys()):
            if smpl.endswith('Full_ROI'):
                smplIdx = -3
            else:
                smplIdx = -2
            prev = smpl.split('_')[smplIdx]
            # prev = smpl.split('_')[-2]
            # if (prev == 'Full'):
            #     prev = smpl.split('_')[-3]
            # print(prev)
            prv2 = "{:03d}".format(int(prev)-2)
            prv = "{:03d}".format(int(prev)-1)
            nxt = "{:03d}".format(int(prev)+1)
            nxt2 = "{:03d}".format(int(prev)+2)
            prv2 = '_'.join(smpl.split('_')[:smplIdx])+'_'+prv2
            prv = '_'.join(smpl.split('_')[:smplIdx])+'_'+prv
            nxt = '_'.join(smpl.split('_')[:smplIdx])+'_'+nxt
            nxt2 = '_'.join(smpl.split('_')[:smplIdx])+'_'+nxt2
            prvSamples = [x for x in sampleInfo.columns if (
                (prv in x) or (prv2 in x))]
            nxtSamples = [x for x in sampleInfo.columns if (
                (nxt in x) or (nxt2 in x))]
            toLocate.append([smpl, prvSamples, nxtSamples])
    # print(toLocate)
    # print('len(toLocate)')
    # print(len(toLocate))
    return(toLocate)

# ToDo : Print plate hints for each well?
# ToDo : Sort samples better for display?
# ToDo : Update dataframes also
# ToDo : Confirm entries with y/n input
# ToDo : Find a more pythonic way to make this code interactive
# ToDo : Fix use of global variables


def enter_locations(toLocate, validWells, AOItoPlateDict, AOItoWellDict,
                    PlateWellDict, wellDFs):
    # ToDo: Need an option to leave input blank to skip sample
    # global AOItoPlateDict
    # global AOItoWellDict
    # global PlateWellDict
    for t in toLocate:
        plate = ''
        well = ''
        # print()
        # print(f'Unlocated well :\t{t[0]}\n')
        # print(t[1])
        # print(t[2])
        if (not((t[1] == []) and (t[2] == []))):
            # Confirm that well has not been located already in a prior run
            try:
                hasPlate = AOItoPlateDict[t[0]]
            except KeyError:
                hasPlate = False
            try:
                hasWell = AOItoWellDict[t[0]]
            except KeyError:
                hasWell = False
            if (hasPlate or hasWell):
                print('has entry')
                time.sleep(1)
                continue
            print('previous:')
            for p in sorted(t[1]):
                try:
                    print('\t\t\t' + p + '\t: ' + str(
                        AOItoPlateDict[p]) + ' ' + AOItoWellDict[p])
                except KeyError:
                    print('\t\t\t' + p + '\t: not in dict')
            print()
            print(f'Unlocated well :\t{t[0]}\n')
            print('next:')
            for n in sorted(t[2]):
                try:
                    print('\t\t\t' + n + '\t: ' + str(
                        AOItoPlateDict[n]) + ' ' + AOItoWellDict[n])
                except KeyError:
                    print('\t\t\t' + n + '\t: not in dict')
            plate = (input(
                'enter plate for this sample, or press enter to pass'))
            if (plate == ''):
                clear_output(wait=False)
                continue
            else:
                try:
                    plate = int(plate)
                except ValueError:
                    print('plate value must be an integer')
                    time.sleep(2)
                    clear_output(wait=False)
                    continue
            well = input('enter well for this sample').upper()
            if (well == ''):
                clear_output(wait=False)
                continue
            clear_output(wait=False)
            print('you entered:')
            print(plate)
            print(well)
            if well in validWells:
                print('\nWell is valid\n\n')
                if (well in PlateWellDict[plate]):
                    print('well already has entry')
                    time.sleep(2)
                else:
                    PlateWellDict[plate].append(well)
                    AOItoWellDict[t[0]] = well
                    AOItoPlateDict[t[0]] = plate
                    # Also add to dataframes
                    wellDFs[plate-1].loc[well[0], well[1:]] = t[0]
            else:
                print('invalid well entered')
                pass
    return (AOItoPlateDict, AOItoWellDict, PlateWellDict, wellDFs)
    # ToDo: Confirm that all entries appear correct (Surface Areas match)


def read_plate_info(masterData, infoPath):
    masterData.sampleInfo = pd.read_csv(infoPath, index_col=0)
    sampleInfo = masterData.sampleInfo
    return (sampleInfo)


def read_config():
    # read in paths from config file
    configDict = {
        'rootDir': '',
        'initialDataPath': '',
        'QCDataPath': '',
        'labWorksheet01Path': '',
        'projectName': '',
        'selectedData': []
    }
    with open('config.txt', 'r') as f:
        lines = f.readlines()
        for line in lines:
            if ((not line.startswith('#')) and (not line.strip() == '')):
                line = line.strip()
                fields = line.split(':')
                print(f'{fields[0]} : {fields[1]}')
                if fields[0].strip() == 'initialDataPath':
                    configDict[fields[0].strip()] = fields[1].strip(
                        ).strip('\'')
                elif fields[0].strip() == 'probeThresholdIdx':
                    configDict[fields[0].strip()] = int(fields[1].strip(
                        ).strip('\''))
                elif fields[0].strip() == 'selectedData':
                    tempList = fields[1].strip().strip('\'').split(',')
                    tempList = [x.strip() for x in tempList]
                    tempList = [x for x in tempList if not x == '']
                    configDict['selectedData'] = tempList
                else:
                    configDict[fields[0].strip()] = fields[1].strip(
                        ).strip('\'')
    # ToDo: Add checks to ensure that minimal fields have been populated.
    # Raise errors or warnings
    return configDict


def get_unique_combos(selectedInfo):
    comboUniques = []
    for c in selectedInfo.columns:
        thisCol = selectedInfo[c]
        combined = '_'.join(thisCol.values)
        comboUniques.append(combined)
    comboUniques = sorted(list(set(comboUniques)))
    print('\nNumber of unique combinations: {}'.format(len(comboUniques)))
    return comboUniques
